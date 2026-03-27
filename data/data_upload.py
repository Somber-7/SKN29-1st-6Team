"""
10.연료별_등록현황 시트 데이터를 TB_FUEL_REG_STAT 테이블에 적재
- 대상 연료: 휘발유(F01), 경유(F02), 전기(F03),
             하이브리드(휘발유+전기)(F04), 하이브리드(경유+전기)(F05)
- 집계 행 (소계/계) 제외
- 하이픈(-) 값 → 0 처리
- ON DUPLICATE KEY UPDATE 로 재실행 안전
"""

import os
import re
import pymysql
import openpyxl

# ──────────────────────────────────────────────
# DB 연결 정보
# ──────────────────────────────────────────────
DB_HOST     = 'localhost'
DB_USER     = 'root'
DB_PASSWORD = 'root1234!'
DB_NAME     = 'car_db'
DB_PORT     = 3306

# ──────────────────────────────────────────────
# 경로
# ──────────────────────────────────────────────
XLSX_DIR    = os.path.join(os.path.dirname(__file__), 'xlsx')
TARGET_SHEET = '10.연료별_등록현황'
BATCH_SIZE  = 1000

# ──────────────────────────────────────────────
# 코드 매핑 (사용 연료 5종만)
# ──────────────────────────────────────────────
FUEL_MAP = {
    '휘발유':              'F01',
    '경유':               'F02',
    '전기':               'F03',
    '하이브리드(휘발유+전기)': 'F04',
    '하이브리드(경유+전기)':  'F05',
}

TYPE_MAP = {
    '승용': 'T01',
    '승합': 'T02',
    '화물': 'T03',
    '특수': 'T04',
}

USAGE_MAP = {
    '비사업용': 'U01',
    '사업용':   'U02',
}

# 열 인덱스(0-based) → 지역코드  (col 3=서울 ~ col 19=제주, col 20=전국계 제외)
REGION_COL_MAP = {
    3:  'R01',   # 서울
    4:  'R02',   # 부산
    5:  'R03',   # 대구
    6:  'R04',   # 인천
    7:  'R05',   # 광주
    8:  'R06',   # 대전
    9:  'R07',   # 울산
    10: 'R08',   # 세종
    11: 'R09',   # 경기
    12: 'R10',   # 강원
    13: 'R11',   # 충북
    14: 'R12',   # 충남
    15: 'R13',   # 전북
    16: 'R14',   # 전남
    17: 'R15',   # 경북
    18: 'R16',   # 경남
    19: 'R17',   # 제주
}


def parse_stat_ym(filename: str) -> str | None:
    """파일명에서 YYYYMM 추출. 예) 2024년_01월_... → '202401'"""
    m = re.search(r'(\d{4})년_(\d{2})월', filename)
    if m:
        return m.group(1) + m.group(2)
    return None


def to_int(value) -> int:
    """셀 값을 정수로 변환. None / 하이픈(-)은 0 반환"""
    if value is None:
        return 0
    s = str(value).strip()
    if s in ('-', '–', '—', ''):
        return 0
    try:
        return int(s.replace(',', ''))
    except ValueError:
        return 0


def parse_sheet(ws, stat_ym: str) -> list[tuple]:
    """
    시트에서 유효 행만 파싱하여 INSERT용 튜플 리스트 반환
    튜플: (STAT_YM, FUEL_CD, TYPE_CD, USAGE_CD, REGION_CD, REG_CNT)
    """
    records = []
    for row in ws.iter_rows(values_only=True):
        fuel_nm  = str(row[0]).strip() if row[0] else ''
        type_nm  = str(row[1]).strip() if row[1] else ''
        usage_nm = str(row[2]).strip() if row[2] else ''

        fuel_cd  = FUEL_MAP.get(fuel_nm)
        type_cd  = TYPE_MAP.get(type_nm)
        usage_cd = USAGE_MAP.get(usage_nm)

        # 대상 연료·차종·용도가 아닌 행 (소계/계/헤더/빈행) 제외
        if not fuel_cd or not type_cd or not usage_cd:
            continue

        for col_idx, region_cd in REGION_COL_MAP.items():
            reg_cnt = to_int(row[col_idx] if col_idx < len(row) else None)
            records.append((stat_ym, fuel_cd, type_cd, usage_cd, region_cd, reg_cnt))

    return records


def batch_upsert(cursor, records: list[tuple]) -> int:
    """1000건 단위 배치 INSERT ... ON DUPLICATE KEY UPDATE"""
    sql = """
        INSERT INTO TB_FUEL_REG_STAT
            (STAT_YM, FUEL_CD, TYPE_CD, USAGE_CD, REGION_CD, REG_CNT)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
            REG_CNT    = VALUES(REG_CNT),
            UPDATED_AT = CURRENT_TIMESTAMP
    """
    total = 0
    for i in range(0, len(records), BATCH_SIZE):
        chunk = records[i:i + BATCH_SIZE]
        cursor.executemany(sql, chunk)
        total += len(chunk)
    return total


def ensure_table(cursor):
    """TB_FUEL_REG_STAT 테이블이 없으면 생성"""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS TB_FUEL_REG_STAT (
            STAT_YM    CHAR(6)     NOT NULL COMMENT '통계년월 (YYYYMM)',
            FUEL_CD    VARCHAR(10) NOT NULL COMMENT '연료코드',
            TYPE_CD    VARCHAR(10) NOT NULL COMMENT '차종코드',
            USAGE_CD   VARCHAR(10) NOT NULL COMMENT '용도코드',
            REGION_CD  VARCHAR(10) NOT NULL COMMENT '지역코드',
            REG_CNT    INT         NOT NULL DEFAULT 0 COMMENT '등록 대수',
            CREATED_AT DATETIME    NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UPDATED_AT DATETIME    NOT NULL DEFAULT CURRENT_TIMESTAMP
                        ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (STAT_YM, FUEL_CD, TYPE_CD, USAGE_CD, REGION_CD),
            INDEX IDX_FUELSTAT_YM     (STAT_YM),
            INDEX IDX_FUELSTAT_FUEL   (FUEL_CD),
            INDEX IDX_FUELSTAT_REGION (REGION_CD)
        ) ENGINE=InnoDB
          DEFAULT CHARSET=utf8mb4
          COLLATE=utf8mb4_unicode_ci
          COMMENT='연료별 차종별 용도별 지역별 등록현황'
    """)


def main():
    # xlsx 파일 목록 수집 (임시 파일 ~$ 제외)
    files = sorted([
        f for f in os.listdir(XLSX_DIR)
        if f.endswith('.xlsx') and not f.startswith('~$')
    ])
    print(f'처리 대상 파일: {len(files)}개\n')

    conn = pymysql.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASSWORD,
        db=DB_NAME, port=DB_PORT, charset='utf8mb4',
        autocommit=False
    )

    try:
        with conn.cursor() as cursor:
            ensure_table(cursor)
            conn.commit()

        total_inserted = 0
        for fname in files:
            stat_ym = parse_stat_ym(fname)
            if not stat_ym:
                print(f'[SKIP] 년월 파싱 실패: {fname}')
                continue

            fpath = os.path.join(XLSX_DIR, fname)
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

            if TARGET_SHEET not in wb.sheetnames:
                wb.close()
                print(f'[SKIP] 시트 없음: {fname}')
                continue

            ws = wb[TARGET_SHEET]
            records = parse_sheet(ws, stat_ym)
            wb.close()

            if not records:
                print(f'[SKIP] 유효 데이터 없음: {fname}')
                continue

            with conn.cursor() as cursor:
                cnt = batch_upsert(cursor, records)
            conn.commit()
            total_inserted += cnt
            print(f'[OK] {fname}  →  {stat_ym}  ({cnt}건 처리)')

        print(f'\n완료: 총 {total_inserted}건 처리')

    except Exception as e:
        conn.rollback()
        print(f'[ERROR] {e}')
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
